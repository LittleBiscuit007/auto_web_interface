from openpyxl import load_workbook


# 加载xlsx文件的位置
workbook = load_workbook("./HS_Redfish_BBFV.xlsx")
# 获取当前活跃的sheet,默认是第一个sheet
booksheet = workbook.active

# print(booksheet)

rows = booksheet.rows
columns = booksheet.columns

# print(rows)
# print(columns)
# 数据开始的行数-1，作为开始遍历的行
i = 11
for row in rows:
    i = i + 1
    line = [col.value for col in row]
    try:
        # 第 i 行 2 列的数据
        cell_data_2 = booksheet.cell(row=i, column=2).value
        # 针对多行合并的数据进行None值填充
        if not cell_data_2:
            j = i-1
            while True:
                if booksheet.cell(row=j, column=2).value:
                    cell_data_2 = booksheet.cell(row=j, column=2).value
                    break
                j = j - 1
        # 第 i 行 3 列的数据
        cell_data_3 = booksheet.cell(row=i, column=3).value.split()
    except Exception:
        break

    print(cell_data_2, cell_data_3)






