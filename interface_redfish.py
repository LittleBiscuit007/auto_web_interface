from openpyxl import load_workbook
import requests
import urllib3
import json

# 关闭ssl认证时，取消警告提示信息
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# 接口测试的根目录和用户名-密码
base_url = "https://10.245.36.148"
username = "root"
password = "0penBmc"


# 加载xlsx文件的位置
workbook = load_workbook("./read_file/HS_Redfish_BBFV.xlsx")
# 获取当前活跃的sheet,默认是第一个sheet
booksheet = workbook.active

# print(booksheet)

rows = booksheet.rows
columns = booksheet.columns

# print(rows)
# print(columns)
# 数据开始的行数-1，作为开始遍历的行
i = 11
for row in rows:
    i = i + 1
    line = [col.value for col in row]
    try:
        # 第 i 行 2 列的数据
        cell_data_2 = booksheet.cell(row=i, column=2).value
        # 针对多行合并的数据进行None值填充
        if not cell_data_2:
            j = i-1
            while True:
                if booksheet.cell(row=j, column=2).value:
                    cell_data_2 = booksheet.cell(row=j, column=2).value
                    break
                j = j - 1
        # 第 i 行 3 列的数据
        cell_data_3 = booksheet.cell(row=i, column=3).value.split()
    except Exception:
        break

    # print(cell_data_2, cell_data_3)
    with open("./interface_redfish.log", 'a+') as log_file:
        # GET
        if cell_data_2 == "GET":
            # headers = {"Content-Type": "application/json"}
            response = requests.get(base_url + cell_data_3[0], auth=(username, password), verify=False)

            # 获取请求的地址和响应的状态码
            get_log = response.url + "\t" + str(response.status_code)
            # 如果请求失败，记录返回的失败信息
            if response.status_code // 100 != 2:
                get_log = get_log + "\t" + response.text + "\n"
            else:
                get_log = get_log + "\n"
            log_file.write(get_log)
        """
        # PATCH
        elif cell_data_2 == "PATCH":
            patch_data = {"Boot": {"BootSourceOverrideEnabled": "Disabled", "BootSourceOverrideTarget": "Cd"},
                         "IndicatorLED": "On"}

            patch_data = json.dumps(patch_data)
            response = requests.patch(base_url + "redfish/v1/Systems/system", data=patch_data, auth=(username, password),
                                      verify=False)
            # 将PATCH请求结果写入log中
            get_log = response.url + "\t" + str(response.status_code) + "\n"
            log_file.write(get_log)

        # POST
        elif cell_data_2 == "POST":
            data_3_list = cell_data_3
            # 组织要添加的数据，格式为字典
            for data_3_str in data_3_list:
                str1_list = data_3_str.split("/")
            
                add_data = {str1_list[-2]: str1_list[-1]}
            
                for i in reversed(str1_list[:-2]):
                    add_data = {i: add_data}
                # print(str1_dict)
                # 要添加的用户的信息
                # add_data = json.dumps({"UserName": "ymh", "Password": "len0vO1"})
                add_data = json.dumps(add_data)
                response = requests.post(base_url + "redfish/v1/AccountService/Accounts", data=add_data,
                                         auth=(username, password), verify=False)
    
                # 获取请求的地址和响应的状态码
                get_log = response.url + "\t" + str(response.status_code)
                # 如果请求失败，记录返回的失败信息
                if response.status_code // 100 != 2:
                    get_log = get_log + "\t" + response.text + "\n"
                else:
                    get_log = get_log + "\n"
                log_file.write(get_log)

        # DELETE
        else:
            # 要删除的用户名
            delete_name = "ymh"
            response = requests.delete(base_url + "redfish/v1/AccountService/Accounts/" + delete_name,
                                       auth=(username, password), verify=False)

            print(response.url, response.status_code)
        """




